"""
「業態転換・名称変更履歴」シートの内容を読み込み、「クリニック一覧」を自動更新するツール。

やること：
  1) 転換前の院（既存行）に業態転換日を記入し、開院フラグを「閉院」にする
  2) 転換後の院がまだクリニック一覧に存在しなければ、新しい行として追加する

このツールは手動で実行してください（自動更新の一部には組み込みません）。
何度実行しても、すでに反映済みの内容は変更されません（安全に再実行できます）。
"""
import sys
from datetime import date
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

BOX_PATH = Path(r"C:\Users\宮城杏奈\Box\総合企画部_特殊案件\その他\院情報一覧カウント\院情報一覧_カウント自動化.xlsx")
LOCAL_PATH = Path.home() / "Documents" / "クリニックDB" / "院情報一覧_カウント自動化.xlsx"
CONVERSION_SHEET = "業態転換・名称変更履歴"
CLINIC_SHEET = "クリニック一覧"


def get_path():
    if BOX_PATH.exists():
        return BOX_PATH
    if LOCAL_PATH.exists():
        return LOCAL_PATH
    raise FileNotFoundError("Excelファイルが見つかりません（Box Driveが起動しているか確認してください）")


def to_int_or_none(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None


def main():
    path = get_path()
    print(f"対象ファイル: {path}")
    wb = openpyxl.load_workbook(path)
    ws_clinic = wb[CLINIC_SHEET]
    ws_conv = wb[CONVERSION_SHEET]

    c_headers = [c.value for c in ws_clinic[1]]
    c_idx = {h: i + 1 for i, h in enumerate(c_headers)}
    required = ["院ID", "正式名称", "TWE表記", "法人名", "開院フラグ", "開院日", "MA日",
                "移転拡張日", "業態転換日", "閉院日", "事業カテゴリ", "ブランド", "業態", "海外／国内"]
    missing = [h for h in required if h not in c_idx]
    if missing:
        print(f"エラー: クリニック一覧に想定した列がありません: {missing}")
        sys.exit(1)

    # 既存行の索引を作る: (院ID, 正式名称) -> 行番号 / 正式名称のみ -> 行番号リスト
    by_id_name = {}
    by_name = {}
    for r in range(2, ws_clinic.max_row + 1):
        name = ws_clinic.cell(row=r, column=c_idx["正式名称"]).value
        if not name:
            continue
        name = str(name).strip()
        cid = to_int_or_none(ws_clinic.cell(row=r, column=c_idx["院ID"]).value)
        by_name.setdefault(name, []).append(r)
        if cid is not None:
            by_id_name[(cid, name)] = r

    conv_headers = [c.value for c in ws_conv[1]]
    conv_idx = {h: i + 1 for i, h in enumerate(conv_headers)}

    updated_before = []
    created_after = []
    skipped_existing = []
    warnings = []
    next_row = ws_clinic.max_row + 1
    today = date.today()

    for r in range(2, ws_conv.max_row + 1):
        def gv(col):
            return ws_conv.cell(row=r, column=conv_idx[col]).value if col in conv_idx else None

        before_id = to_int_or_none(gv("転換前院ID"))
        before_name = str(gv("転換前名称") or "").strip()
        conv_date = gv("業態転換日")
        after_id = to_int_or_none(gv("転換後院ID"))
        after_name = str(gv("転換後名称") or "").strip()

        if not before_name or conv_date is None:
            continue
        conv_date_only = conv_date.date() if hasattr(conv_date, "date") else conv_date

        # ── ① 転換前の院を更新 ──
        before_row = by_id_name.get((before_id, before_name)) if before_id is not None else None
        if before_row is None:
            before_row = next(iter(by_name.get(before_name, [])), None)
        if before_row is None:
            warnings.append(f"No.{r-1}: 転換前の院「{before_name}」がクリニック一覧に見つかりません")
        else:
            changed = []
            cur_conv = ws_clinic.cell(row=before_row, column=c_idx["業態転換日"]).value
            cur_conv_only = cur_conv.date() if hasattr(cur_conv, "date") else cur_conv
            if cur_conv_only != conv_date_only:
                ws_clinic.cell(row=before_row, column=c_idx["業態転換日"], value=conv_date)
                changed.append("業態転換日")
            # 業態転換日が未来の場合、転換前の院はまだ営業中なので開院フラグは変更しない
            cur_flag = str(ws_clinic.cell(row=before_row, column=c_idx["開院フラグ"]).value or "").strip()
            if conv_date_only <= today and cur_flag != "閉院":
                ws_clinic.cell(row=before_row, column=c_idx["開院フラグ"], value="閉院")
                changed.append("開院フラグ")
            if changed:
                updated_before.append(f"No.{r-1}: {before_name}（{'・'.join(changed)}を更新）")

        # ── ② 転換後の院を確認・追加 ──
        if not after_name:
            continue
        exists = False
        conflict = False
        if after_id is not None:
            if (after_id, after_name) in by_id_name:
                exists = True
            elif after_name in by_name:
                conflict = True
        else:
            if after_name in by_name:
                exists = True

        if conflict:
            warnings.append(f"No.{r-1}: 転換後の院「{after_name}」院ID{after_id}が既存データと食い違うため、自動追加をスキップしました（要確認）")
            continue
        if exists:
            skipped_existing.append(f"No.{r-1}: {after_name}（既存）")
            continue

        # 新規行を作成
        flag = "開院" if conv_date_only <= today else "開院前"
        row_vals = {
            "院ID": after_id if after_id is not None else None,
            "正式名称": after_name,
            "TWE表記": None,
            "法人名": gv("転換後法人名"),
            "開院フラグ": flag,
            "開院日": conv_date,
            "MA日": None,
            "移転拡張日": None,
            "業態転換日": None,
            "閉院日": None,
            "事業カテゴリ": gv("転換後事業カテゴリ"),
            "ブランド": gv("転換後ブランド"),
            "業態": gv("転換後業態"),
            "海外／国内": gv("転換後海外／国内"),
        }
        for h, v in row_vals.items():
            ws_clinic.cell(row=next_row, column=c_idx[h], value=v)
        by_name.setdefault(after_name, []).append(next_row)
        if after_id is not None:
            by_id_name[(after_id, after_name)] = next_row
        created_after.append(f"No.{r-1}: {after_name}（新規追加、行{next_row}）")
        next_row += 1

    wb.save(path)

    print("\n=== 反映結果 ===")
    print(f"転換前の院を更新: {len(updated_before)}件")
    for line in updated_before:
        print(f"  ・{line}")
    print(f"転換後の院を新規追加: {len(created_after)}件")
    for line in created_after:
        print(f"  ・{line}")
    print(f"すでに反映済みでスキップ: {len(skipped_existing)}件")
    if warnings:
        print(f"\n⚠ 要確認: {len(warnings)}件")
        for line in warnings:
            print(f"  ・{line}")
    print("\n完了しました。")


if __name__ == "__main__":
    main()
